#!/usr/bin/env python
# coding: utf-8

# In[5]:


import pandas as pd
import numpy as np
import hashlib
import openpyxl
from openpyxl.styles import PatternFill
import datetime
from datetime import date
from datetime import datetime
from datetime import timedelta
import math
from math import sqrt
import ipywidgets as widgets
from IPython.display import display
import geopandas as gpd
from pyproj import CRS
import pyproj
from shapely.geometry import Point
from geopy.distance import geodesic
import xlsxwriter

file_path = r"F:\Temp\Automation\Arpita\automation\streamlit\Input\Rajathan_GT_January_2023.xlsx" # Replace 'filename.txt' with the actual file name/path
output_file_path = r"F:\Temp\Automation\Arpita\automation\streamlit\OUTPUT\test_1.2.xlsx"


dataframe = pd.read_excel(file_path)

df = dataframe[["REFID", "DATE", "DISTRICT", "TALUKA", "EMPLOYEE NAME", "COORDINATES"]]

# Remove extra white spaces from every column
df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
print(df)

# Convert the "COORDINATES" column to string
df['COORDINATES'] = df['COORDINATES'].astype(str)

# Extract latitude and longitude from the "COORDINATES" column
df[['LATITUDE', 'LONGITUDE']] = df['COORDINATES'].str.split(',', expand=True)

# Convert latitude and longitude columns to numeric (optional, if needed)
df['LATITUDE'] = pd.to_numeric(df['LATITUDE'])
df['LONGITUDE'] = pd.to_numeric(df['LONGITUDE'])



# Replace "GMT+05:30" and "IST" with blank from the "DATE" column
df['DATE'] = df['DATE'].str.replace('GMT\+05:30|GMT\+01:00|IST', '', regex=True)
df['DATE'] = df['DATE'].str.strip()
df['DATE'] = df['DATE'].str.replace('  -', '', regex=False)

# Convert "DATE" column to datetime
df['DATE'] = pd.to_datetime(df['DATE'], format='%Y.%m.%d %H:%M:%S')

# Define a function to calculate the week number based on date ranges
def calculate_week_number(date):
    if 1 <= date.day <= 7:
        return f"{date.strftime('%b')} 01 week"
    elif 8 <= date.day <= 14:
        return f"{date.strftime('%b')} 02 week"
    elif 15 <= date.day <= 21:
        return f"{date.strftime('%b')} 03 week"
    else:
        return f"{date.strftime('%b')} 04 week"

# Apply the function to create the "Week Number" column
df["Week Number"] = df["DATE"].apply(calculate_week_number)

# Extract date
df["Date"] = df["DATE"].dt.strftime('%d-%m-%Y')

# Extract time
df["Time"] = df["DATE"].dt.strftime('%I:%M:%S %p')
# Convert the time column to the desired format
# df['Time'] = pd.to_datetime(df['Time'], format='%I:%M:%S %p').dt.time

# Create the "Join" column with the desired joined text
df["Join"] = df["DISTRICT"] + "_" + df["TALUKA"] + "_" + df["EMPLOYEE NAME"] + "_" + df["Date"]

# Sort the DataFrame by the "Join" column in ascending order
df = df.sort_values(by='Join', ascending=True)

# Calculate GT Count
df["GT Done"] = df.groupby(["Week Number", "Join"])["EMPLOYEE NAME"].transform("count")

# long / lat columns to geodataframe geomtry all other columns attributes
gdf = gpd.GeoDataFrame(
    geometry=gpd.points_from_xy(df.LONGITUDE, df.LATITUDE, crs="EPSG:4326"), data=df)

# Reporject to UTM
gdf_utm = gdf.to_crs('EPSG:32643')

gdf_utm = gdf_utm.reset_index()

# Sort the DataFrame by the "Join" column in ascending order
gdf_utm = gdf_utm.sort_values(by='Join', ascending=True)

gdf_utm['Join_check'] = gdf_utm['Join'].shift(-1)
gdf_utm['point2'] = gdf_utm['geometry'].shift(-1)
gdf_utm['REFID2'] = gdf_utm['REFID'].shift(-1)

gdf_utm['Distance'] = np.nan  # Initialize the "Distance" column with NaN values
mask = (gdf_utm['Join'] == gdf_utm['Join_check'])
gdf_utm_masked = gdf_utm[mask]
gdf_utm.loc[mask, 'Distance'] = gdf_utm_masked['geometry'].distance(gdf_utm_masked['point2'])/ 1000
gdf_utm = gdf_utm.drop(columns=['DATE', 'point2'])



# Calculate average distance for each join
average_distance = gdf_utm.groupby('Join')['Distance'].mean()

# Merge average distances with the original DataFrame
df_with_avg_distance = pd.merge(gdf_utm, average_distance, on='Join', suffixes=('', '_avg'))

# Rename the average distance column
df_with_avg_distance = df_with_avg_distance.rename(columns={'Distance_avg': 'GT to GT Average Distance'})

# Convert the "Time" column to the desired format
#df_with_avg_distance['Time'] = pd.to_datetime(df_with_avg_distance['Time'], format='%I:%M:%S %p').dt.time

df_with_avg_distance['Time'] = pd.to_datetime(df_with_avg_distance['Time'], format='%I:%M:%S %p')
df_with_avg_distance.sort_values(by=['Join', 'Time'], ascending=[True, True], inplace=True)

# Shift the "Join" and "Time" columns by one row
df_with_avg_distance['Join_next'] = df_with_avg_distance['Join'].shift(-1)
df_with_avg_distance['Time_next'] = df_with_avg_distance['Time'].shift(-1)

df_with_avg_distance['Time Difference'] = df_with_avg_distance.apply(
    lambda row: row['Time_next'] - row['Time'] if row['Join_next'] == row['Join'] else '',
    axis=1)


# Convert the time difference to the desired format
df_with_avg_distance['Time Difference'] = df_with_avg_distance['Time Difference'].dt.total_seconds().fillna(0).astype(int).apply(
    lambda seconds: str(timedelta(seconds=seconds)).split('.', 2)[0])


# Remove the additional columns
df_with_avg_distance = df_with_avg_distance.drop(columns=['Join_next', 'Time_next'])

# Convert the "Time Difference" column to timedelta format
df_with_avg_distance['Time Difference'] = pd.to_timedelta(df_with_avg_distance['Time Difference'])

# Calculate average time difference
df_with_avg_distance['GT to GT Average Time'] = df_with_avg_distance.groupby('Join')['Time Difference'].transform('mean')


# Convert the "Average Time Difference" column to the desired format (HH:MM:SS)
df_with_avg_distance['GT to GT Average Time'] = df_with_avg_distance['GT to GT Average Time'].dt.total_seconds().fillna(0).astype(int)
df_with_avg_distance['GT to GT Average Time'] = df_with_avg_distance['GT to GT Average Time'].apply(lambda x: f"{int(x//3600):02d}:{int((x//60)%60):02d}:{int(x%60):02d}")

# Use numpy.where to apply the IFERROR logic
df_with_avg_distance['GT to GT Average Time'] = np.where(df_with_avg_distance['GT to GT Average Time'].str.startswith('0 days'), '', df_with_avg_distance['GT to GT Average Time'])


# Save the updated DataFrame with time difference to the output file
df_with_avg_distance.to_excel(output_file_path, index=False)

# Create the sidebar widgets for District and Taluka selection
district_dropdown = widgets.Dropdown(
    options=df_with_avg_distance['DISTRICT'].unique(),
    description='District:'
)
taluka_dropdown = widgets.Dropdown(description='Taluka:')

# Update the Taluka dropdown options based on the selected District
def update_taluka_options(*args):
    selected_district = district_dropdown.value
    taluka_dropdown.options = df_with_avg_distance[df_with_avg_distance['DISTRICT'] == selected_district]['TALUKA'].unique()

# Register the update_taluka_options function as the event handler
district_dropdown.observe(update_taluka_options, 'value')



def create_pivot_table(district, taluka):
    # Filter the DataFrame based on the selected District and Taluka
    filtered_df = df_with_avg_distance[(df_with_avg_distance['DISTRICT'] == district) & (df_with_avg_distance['TALUKA'] == taluka)]


    # Create the pivot table with Employee Name as row labels, Week Number as column labels,
    # and aggregate values using count, mean, and first functions
    pivot_table = pd.pivot_table(filtered_df, values=['GT Done', 'GT to GT Average Distance', 'GT to GT Average Time'],
                                 index='EMPLOYEE NAME', columns='Week Number',
                                 aggfunc={'GT Done': 'count', 'GT to GT Average Distance': 'mean',
                                          'GT to GT Average Time': 'first'})
    
    # Add District and Taluka as headers in the pivot table
    pivot_table = pd.concat([pivot_table], keys=[f'District: {district}, Taluka: {taluka}'])
    
    # Format the pivot table
    formatted_table = pivot_table.round(3).fillna('')


    # Add Grand Total row
    # formatted_table.loc['Grand Total'] = formatted_table.sum()

    # Get the unique weeks from the "Week Number" column
    unique_weeks = formatted_table.columns.levels[1].unique()

    # Reorder the columns based on the unique weeks
    reordered_columns = []
    for week in unique_weeks:
        reordered_columns.extend([(col[0], col[1]) for col in formatted_table.columns if col[1] == week])

    formatted_table = formatted_table[reordered_columns]
    
    # Set up the Styler object for customizing the pivot table display
    styler = formatted_table.style

    # Set the header style for District and Taluka
    styler = styler.set_table_styles([
        {'selector': 'th', 'props': [('text-align', 'left'), ('font-weight', 'bold'), ('border', '1px solid black')]}
    ])

    # Set the styles for the Week Number and Employee Name columns
    styler = styler.set_properties(subset=pd.IndexSlice[:, :], **{'text-align': 'center', 'font-weight': 'bold',
                                                                    'border': '1px solid black'})

    # Set the styles for the other columns (GT Done, GT to GT Average Distance, GT to GT Average Time)
    styler = styler.format({'GT Done': '{:.0f}', 'GT to GT Average Distance': '{:.2f}', 'GT to GT Average Time': '{:.2f}'})
    styler = styler.set_table_styles([
        {'selector': 'td', 'props': [('border', '1px solid black')]}
    ])

    # Set the row height and column width to a fixed value
    styler = styler.set_properties(subset=pd.IndexSlice[:, :], **{'height': '25px', 'width': '120px'})

    # Save the pivot table to an Excel file
    styler.to_excel(r"F:\Temp\Automation\Arpita\automation\streamlit\OUTPUT\pivot_table_A.xlsx")

    # Display a success message
    print("Pivot table saved successfully.")

# Create the submit button and its click event handler
button = widgets.Button(description='Submit')

def on_button_click(b):
    district = district_dropdown.value
    taluka = taluka_dropdown.value
    create_pivot_table(district, taluka)

button.on_click(on_button_click)

# Display the widgets
display(district_dropdown, taluka_dropdown, button)






# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[1]:





#  # patternfill
#     writer = pd.ExcelWriter(r"F:\Temp\Automation\Arpita\automation\streamlit\OUTPUT\pivot_table.xlsx", engine='openpyxl')
# 
#     # Write the pivot table to the Excel file
#     pivot_table.to_excel(writer, sheet_name='Pivot Table', index=True)
# 
#     # Get the workbook and the worksheet
#     workbook = writer.book
#     worksheet = workbook['Pivot Table']
# 
#     # Define the red fill pattern
#     red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
# 
#     # Iterate through the cells in the "GT Done" column and apply red fill to cells with values less than 50
#     for cell in worksheet['C']:
#         if cell.row != 1 and cell.value and int(cell.value) < 50:
#             cell.fill = red_fill
#     # Save the Excel file
#     writer.save()

# In[ ]:


def create_pivot_table(district, taluka):
    # Filter the DataFrame based on the selected District and Taluka
    filtered_df = df_with_avg_distance[(df_with_avg_distance['DISTRICT'] == district) & (df_with_avg_distance['TALUKA'] == taluka)]


    # Create the pivot table with Employee Name as row labels, Week Number as column labels,
    # and aggregate values using count, mean, and first functions
    pivot_table = pd.pivot_table(filtered_df, values=['GT Done', 'GT to GT Average Distance', 'GT to GT Average Time'],
                                 index='EMPLOYEE NAME', columns='Week Number',
                                 aggfunc={'GT Done': 'count', 'GT to GT Average Distance': 'mean',
                                          'GT to GT Average Time': 'first'})

    # Format the pivot table
    formatted_table = pivot_table.round(3).fillna('')

    # Add Grand Total row
    # formatted_table.loc['Grand Total'] = formatted_table.sum()

    # Get the unique weeks from the "Week Number" column
    unique_weeks = formatted_table.columns.levels[1].unique()

    # Reorder the columns based on the unique weeks
    reordered_columns = []
    for week in unique_weeks:
        reordered_columns.extend([(col[0], col[1]) for col in formatted_table.columns if col[1] == week])

    formatted_table = formatted_table[reordered_columns]
    
        # Save the pivot table to an Excel file using xlsxwriter for formatting
    workbook = xlsxwriter.Workbook(r"F:\Temp\Automation\Arpita\automation\streamlit\OUTPUT\pivot_table_A.xlsx")
    worksheet = workbook.add_worksheet()

    # Set the header style for District and Taluka
    header_format = workbook.add_format({'text_wrap': True, 'align': 'left', 'bold': True, 'border': 1})
    worksheet.set_row(0, None, header_format)
    worksheet.set_column(1, len(formatted_table.columns), 15, header_format)

    # Set the styles for the Week Number and Employee Name columns
    cell_format = workbook.add_format({'align': 'center', 'bold': True, 'border': 1})
    worksheet.set_column(1, len(formatted_table.columns), None, cell_format)

    # Set the styles for the other columns (GT Done, GT to GT Average Distance, GT to GT Average Time)
    data_format = workbook.add_format({'border': 1})
    for col in range(len(formatted_table.columns)):
        worksheet.set_column(col + 1, col + 1, None, data_format)

    # Write the pivot table data to the worksheet
    for row_num, (index, data) in enumerate(formatted_table.iterrows(), start=1):
        worksheet.write(row_num, 0, index)
        for col_num, value in enumerate(data.values, start=1):
            worksheet.write(row_num, col_num, value)

    # Write the row labels (Employee Names)
    worksheet.write_column(1, 0, formatted_table.index.tolist())

    # Write the column labels (Week Numbers)
    worksheet.write_row(0, 1, formatted_table.columns.get_level_values(1).tolist())

    # Close the workbook
    workbook.close()

    # Display a success message
    print("Pivot table saved successfully.")


# In[ ]:





# In[ ]:






